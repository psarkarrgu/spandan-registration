def confirm_dialog(title, message, confirm_button_text="Confirm"):
    """Display a confirmation dialog."""
    return st.warning(f"{title}: {message}", icon="⚠️")
"""
Utility functions for the Event Registration System.
"""
import streamlit as st
import pandas as pd
import io
import base64
import os
import datetime
import config
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import plotly.express as px
import plotly.graph_objects as go
from database import Database

def create_registration_template():
    """Create and return a registration template Excel file."""
    if not os.path.exists(config.TEMPLATE_DIR):
        os.makedirs(config.TEMPLATE_DIR)
    
    template_path = os.path.join(config.TEMPLATE_DIR, "registration_template.xlsx")
    
    # Create a new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Registration"
    
    # Define headers
    headers = ["Name", "Email", "Phone", "College", "Group Name", "Event ID"]
    
    # Add headers to the worksheet
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    # Add some example data
    example_data = [
        ["John Doe", "john@example.com", "9876543210", "ABC College", "Team Alpha", "1"],
        ["Jane Smith", "jane@example.com", "8765432109", "XYZ University", "Team Alpha", "1"],
    ]
    
    for row_num, data_row in enumerate(example_data, 2):
        for col_num, value in enumerate(data_row, 1):
            ws.cell(row=row_num, column=col_num).value = value
    
    # Add a note about required fields
    ws.cell(row=len(example_data) + 3, column=1).value = "* Required fields"
    ws.cell(row=len(example_data) + 4, column=1).value = "Note: Event ID refers to the ID of the event in the system."
    
    # Set column widths
    for col_num in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 20
    
    # Save the template
    wb.save(template_path)
    
    return template_path

def get_download_link(file_path, link_text):
    """Generate a download link for a file."""
    with open(file_path, 'rb') as f:
        data = f.read()
    
    b64 = base64.b64encode(data).decode()
    file_name = os.path.basename(file_path)
    mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    href = f'<a href="data:{mime_type};base64,{b64}" download="{file_name}">{link_text}</a>'
    
    return href

def dataframe_to_excel(df):
    """Convert a pandas DataFrame to an Excel file in memory."""
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name="Data")
    
    return output.getvalue()

def validate_uploaded_data(df, events):
    """Validate uploaded registration data."""
    errors = []
    warnings = []
    
    # Check required columns
    required_columns = ['Name', 'Event ID']
    for col in required_columns:
        if col not in df.columns:
            errors.append(f"Missing required column: {col}")
    
    # If there are column errors, return early
    if errors:
        return False, errors, warnings
    
    # Get valid event IDs
    valid_event_ids = [str(event['id']) for event in events]
    
    # Validate rows
    for idx, row in df.iterrows():
        # Check for missing required fields
        if pd.isna(row['Name']) or str(row['Name']).strip() == '':
            errors.append(f"Row {idx+2}: Missing required field 'Name'")
        
        if pd.isna(row['Event ID']) or str(row['Event ID']).strip() == '':
            errors.append(f"Row {idx+2}: Missing required field 'Event ID'")
        elif str(int(row['Event ID'])) not in valid_event_ids:
            errors.append(f"Row {idx+2}: Invalid Event ID '{row['Event ID']}'. Valid IDs are: {', '.join(valid_event_ids)}")
        
        # Check for valid email format if provided
        if 'Email' in df.columns and not pd.isna(row['Email']) and row['Email'].strip() != '':
            if '@' not in row['Email'] or '.' not in row['Email']:
                warnings.append(f"Row {idx+2}: Email '{row['Email']}' may not be valid")
        
        # Check for valid phone format if provided
        if 'Phone' in df.columns and not pd.isna(row['Phone']) and str(row['Phone']).strip() != '':
            phone_str = str(row['Phone']).strip()
            if not phone_str.isdigit() or len(phone_str) < 10:
                warnings.append(f"Row {idx+2}: Phone number '{phone_str}' may not be valid")
    
    return len(errors) == 0, errors, warnings

def prepare_data_for_db(df):
    """Prepare uploaded data for database insertion."""
    # Make a copy to avoid modifying the original DataFrame
    df_copy = df.copy()
    
    # Standardize column names: convert to lowercase and replace spaces with underscores
    df_copy.columns = [col.lower().replace(' ', '_') for col in df_copy.columns]
    
    # Check if 'event_id' exists, if not, try to find 'event id' or similar
    if 'event_id' not in df_copy.columns:
        # Try to find the event ID column from the original DataFrame
        if 'Event ID' in df.columns:
            df_copy['event_id'] = df['Event ID']
        elif 'event id' in df_copy.columns:
            df_copy['event_id'] = df_copy['event id']
            df_copy = df_copy.drop(columns=['event id'])
        else:
            # Check for any column containing 'event' and 'id'
            possible_columns = [col for col in df_copy.columns if 'event' in col.lower() and 'id' in col.lower()]
            if possible_columns:
                df_copy['event_id'] = df_copy[possible_columns[0]]
                df_copy = df_copy.drop(columns=[possible_columns[0]])
    
    # Ensure event_id is an integer
    df_copy['event_id'] = df_copy['event_id'].astype(int)
    
    # Make sure all required columns exist
    required_columns = ['name', 'email', 'phone', 'college', 'group_name', 'event_id']
    for col in required_columns:
        if col not in df_copy.columns:
            # Create empty column if missing
            df_copy[col] = None
    
    # Convert to list of dictionaries
    return df_copy.to_dict('records')

def create_chart(chart_type, data, title, x_label, y_label):
    """Create a chart using Plotly."""
    if chart_type == 'bar':
        fig = px.bar(data, x=x_label, y=y_label, title=title)
    elif chart_type == 'pie':
        fig = px.pie(data, names=x_label, values=y_label, title=title)
    elif chart_type == 'line':
        fig = px.line(data, x=x_label, y=y_label, title=title)
    else:
        fig = go.Figure()
        fig.update_layout(title=title)
    
    fig.update_layout(
        title_font=dict(size=24),
        xaxis_title=x_label,
        yaxis_title=y_label,
        template="plotly_white"
    )
    
    return fig

def apply_custom_css():
    """Apply custom CSS to the Streamlit app."""
    css = """
    <style>
        .stApp {
            
            margin: 0 auto;
        }
          
        
        div[data-testid="stToolbar"] {
            visibility: hidden;
        }
        
        
    </style>
    """
    st.markdown(css, unsafe_allow_html=True)
    
    # Add some JavaScript for camera functionality
    camera_js = """
    <script>
    function captureIdCard() {
        if (!navigator.mediaDevices || !navigator.mediaDevices.getUserMedia) {
            alert('Your browser does not support camera access!');
            return;
        }
        
        navigator.mediaDevices.getUserMedia({ video: true })
            .then(function(stream) {
                // Camera access successful
                const video = document.createElement('video');
                video.srcObject = stream;
                video.play();
                
                // You would normally implement capture logic here
                // This is just a placeholder
            })
            .catch(function(error) {
                alert('Error accessing camera: ' + error.message);
            });
    }
    </script>
    """
    st.markdown(camera_js, unsafe_allow_html=True)

def resize_image(image_bytes, max_size_kb=500):
    """Resize and compress an image to a smaller size for storage with better quality preservation."""
    try:
        from PIL import Image, ImageEnhance
        import io
        import numpy as np
        
        # Open the image
        image = Image.open(io.BytesIO(image_bytes))
        
        # Convert to RGB if needed (to handle PNG with alpha channel)
        if image.mode == 'RGBA':
            # Create a white background
            background = Image.new('RGB', image.size, (255, 255, 255))
            # Composite the image with the white background
            image = Image.alpha_composite(background.convert('RGBA'), image)
            image = image.convert('RGB')
        elif image.mode != 'RGB':
            image = image.convert('RGB')
        
        # Original dimensions
        original_width, original_height = image.size
        print(f"DEBUG: Original dimensions: {original_width}x{original_height}")
        
        # Calculate target dimensions while maintaining aspect ratio
        max_dimension = 1200  # Higher max dimension for better quality
        
        if original_width > max_dimension or original_height > max_dimension:
            if original_width > original_height:
                new_width = max_dimension
                new_height = int(original_height * (max_dimension / original_width))
            else:
                new_height = max_dimension
                new_width = int(original_width * (max_dimension / original_height))
            
            # Use LANCZOS resampling for better quality
            image = image.resize((new_width, new_height), Image.LANCZOS)
            print(f"DEBUG: Resized to {new_width}x{new_height}")
        
        # Enhance the image slightly
        enhancer = ImageEnhance.Sharpness(image)
        image = enhancer.enhance(1.2)  # Slightly sharpen
        
        # Convert to NumPy array for processing
        img_array = np.array(image)
        
        # Apply moderate noise reduction if image is noisy
        # (This is a simple approach - OpenCV would offer more sophisticated methods)
        if np.std(img_array) > 50:  # Crude noise detection
            from scipy import ndimage
            img_array = ndimage.gaussian_filter(img_array, sigma=0.5)
            image = Image.fromarray(img_array.astype('uint8'))
            print("DEBUG: Applied light noise reduction")
        
        # Try different quality settings to find optimal compression
        output = io.BytesIO()
        
        # Start with high quality and decrease until size constraint is met
        quality = 95
        image.save(output, format='JPEG', quality=quality, optimize=True)
        
        # Try different quality settings to meet size constraint
        while output.tell() > max_size_kb * 1024 and quality > 65:
            output = io.BytesIO()
            quality -= 5
            print(f"DEBUG: Reducing JPEG quality to {quality}")
            image.save(output, format='JPEG', quality=quality, optimize=True)
        
        result = output.getvalue()
        print(f"DEBUG: Original size: {len(image_bytes)/1024:.1f}KB, New size: {len(result)/1024:.1f}KB")
        return result
    except Exception as e:
        print(f"ERROR in resize_image: {str(e)}")
        # If optimization fails, return original if under max size or truncated version
        if len(image_bytes) <= max_size_kb * 1024:
            return image_bytes
        return image_bytes[:max_size_kb * 1024]  # Truncate to max size

def export_to_csv(df, filename="export.csv"):
    """Generate a download link for a DataFrame as CSV."""
    csv = df.to_csv(index=False)
    b64 = base64.b64encode(csv.encode()).decode()
    href = f'<a href="data:file/csv;base64,{b64}" download="{filename}">Download CSV</a>'
    return href

def set_theme(is_dark_mode):
    """Set the theme (light/dark) for the app."""
    if is_dark_mode:
        st.markdown('<body class="dark-mode">', unsafe_allow_html=True)
    else:
        st.markdown('<body>', unsafe_allow_html=True)

def export_to_excel(df, filename="export.xlsx"):
    """Generate a download link for a DataFrame as Excel."""
    output = io.BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')
    df.to_excel(writer, index=False, sheet_name='Sheet1')
    writer.close()
    
    b64 = base64.b64encode(output.getvalue()).decode()
    href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{filename}">Download Excel</a>'
    return href

def format_check_in_time(time_str):
    """Format check-in time for display."""
    if not time_str:
        return "-"
    
    try:
        dt = datetime.datetime.strptime(time_str, "%Y-%m-%d %H:%M:%S")
        return dt.strftime("%d %b %Y, %I:%M %p")
    except:
        return time_str

def show_loading_spinner(text="Loading..."):
    """Display a loading spinner with text."""
    with st.spinner(text):
        placeholder = st.empty()
        return placeholder

def get_color_scale():
    """Get a nice color scale for charts."""
    return px.colors.sequential.Blues